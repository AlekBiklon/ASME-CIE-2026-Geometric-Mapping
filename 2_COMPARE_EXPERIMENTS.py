# ============================================================
# SCRIPT NAME:
# 2_COMPARE_EXPERIMENTS.py
#
# ASME CIE 2026 STUDENT HACKATHON
#
# UNIVERSAL FINAL EXPERIMENT COMPARISON
#
# PURPOSE:
#
#   Automatically discover experiments:
#
#       B01
#       B02
#       B03
#       ...
#       B10
#       B100
#
#   Compare:
#
#       GEOMETRIC_MAPPING
#       vs
#       AUTODESK neuralCAD-Edit
#
#
# IMPORTANT:
#
#   This script DOES NOT:
#
#       - run neuralCAD-Edit
#       - call OpenAI API
#       - modify CAD geometry
#       - run 1_MAIN.py
#
#   It ONLY reads existing experiment reports.
#
#
# OUTPUT:
#
#   comparison/
#       ASME_method_comparison.xlsx
#
#
# EXCEL SHEETS:
#
#   1. Experiment Comparison
#   2. Method Summary
#   3. Quality Metrics
#   4. Adaptive Details
#   5. Data Availability
#   6. Experiment Index
#
# ============================================================


import json
import math
import re
from pathlib import Path
from statistics import mean, median

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent

EXPERIMENTS_DIR = (
    PROJECT_ROOT
    / "experiments"
)

COMPARISON_DIR = (
    PROJECT_ROOT
    / "comparison"
)

OUTPUT_XLSX = (
    COMPARISON_DIR
    / "ASME_method_comparison.xlsx"
)


# ============================================================
# CONSOLE
# ============================================================

def header(title):

    print("\n" + "=" * 100)
    print(title)
    print("=" * 100)


# ============================================================
# SAFE JSON
# ============================================================

def load_json(path):

    if path is None:
        return None

    try:
        path = Path(path)

    except Exception:
        return None

    if not path.exists():
        return None

    if not path.is_file():
        return None

    try:

        with open(
            path,
            "r",
            encoding="utf-8"
        ) as f:

            return json.load(f)

    except Exception as exc:

        print("\nWARNING — JSON READ FAILED")
        print(path)
        print(type(exc).__name__, ":", exc)

        return None


# ============================================================
# DICTIONARY HELPERS
# ============================================================

def nested_get(
    data,
    *keys,
    default=None
):

    current = data

    for key in keys:

        if not isinstance(current, dict):
            return default

        if key not in current:
            return default

        current = current[key]

    return current


def first_not_none(*values):

    for value in values:

        if value is not None:
            return value

    return None


# ============================================================
# RECURSIVE SEARCH
# ============================================================

def recursive_find(
    data,
    target_keys
):

    if data is None:
        return None

    if isinstance(target_keys, str):
        target_keys = {target_keys}

    else:
        target_keys = set(target_keys)

    if isinstance(data, dict):

        for key, value in data.items():

            if (
                key in target_keys
                and
                value is not None
            ):
                return value

        for value in data.values():

            result = recursive_find(
                value,
                target_keys
            )

            if result is not None:
                return result

    elif isinstance(data, list):

        for item in data:

            result = recursive_find(
                item,
                target_keys
            )

            if result is not None:
                return result

    return None


# ============================================================
# SAFE TYPES
# ============================================================

def safe_float(value):

    try:

        if value is None:
            return None

        value = float(value)

        if not math.isfinite(value):
            return None

        return value

    except Exception:
        return None


def safe_int(value):

    try:

        if value is None:
            return None

        return int(value)

    except Exception:
        return None


def safe_bool(value):

    if isinstance(value, bool):
        return value

    if value is None:
        return None

    if isinstance(value, (int, float)):

        if value == 1:
            return True

        if value == 0:
            return False

    if isinstance(value, str):

        value = value.strip().lower()

        if value in {
            "true",
            "yes",
            "success",
            "successful",
            "1"
        }:
            return True

        if value in {
            "false",
            "no",
            "failed",
            "failure",
            "0"
        }:
            return False

    return None


def excel_bool(value):

    if value is True:
        return "YES"

    if value is False:
        return "NO"

    return "N/A"


# ============================================================
# EXPERIMENT DISCOVERY
# ============================================================

def discover_experiments():

    if not EXPERIMENTS_DIR.exists():

        raise FileNotFoundError(
            f"Experiments directory not found:\n"
            f"{EXPERIMENTS_DIR}"
        )

    pattern = re.compile(
        r"^B(\d+)$",
        re.IGNORECASE
    )

    experiments = []

    for folder in EXPERIMENTS_DIR.iterdir():

        if not folder.is_dir():
            continue

        match = pattern.match(
            folder.name
        )

        if not match:
            continue

        number = int(
            match.group(1)
        )

        experiments.append(
            (
                number,
                folder.name,
                folder
            )
        )

    experiments.sort(
        key=lambda item: item[0]
    )

    return experiments


# ============================================================
# FILE HELPERS
# ============================================================

def first_existing(candidates):

    for candidate in candidates:

        if candidate is None:
            continue

        candidate = Path(candidate)

        if candidate.exists():
            return candidate

    return None


def find_latest_file(
    directory,
    patterns
):

    directory = Path(directory)

    if not directory.exists():
        return None

    files = []

    for pattern in patterns:

        files.extend(
            directory.rglob(pattern)
        )

    if not files:
        return None

    files.sort(
        key=lambda path:
            path.stat().st_mtime,
        reverse=True
    )

    return files[0]


# ============================================================
# PATH DISCOVERY
# ============================================================

def discover_paths(
    experiment_id,
    experiment_dir
):

    experiment_dir = Path(
        experiment_dir
    )

    # --------------------------------------------------------
    # INPUT
    # --------------------------------------------------------

    edit_request = first_existing([

        experiment_dir
        / "input"
        / "edit_request.json"
    ])

    # --------------------------------------------------------
    # OUR METHOD
    # --------------------------------------------------------

    our_metrics = first_existing([

        experiment_dir
        / "metrics"
        / "metrics_report.json",

        experiment_dir
        / "our_method_output"
        / "metrics_report.json"
    ])

    our_edit = first_existing([

        experiment_dir
        / "our_method_output"
        / f"{experiment_id}_edit_report.json",

        experiment_dir
        / "our_method_output"
        / "edit_report.json"
    ])

    our_module_edit = first_existing([

        experiment_dir
        / "our_method_output"
        / "cad_edit_report.json"
    ])

    our_step = first_existing([

        experiment_dir
        / "our_method_output"
        / f"{experiment_id}_result.step",

        experiment_dir
        / "our_method_output"
        / f"{experiment_id}_result.stp"
    ])

    # --------------------------------------------------------
    # GT
    # --------------------------------------------------------

    ground_truth = first_existing([

        experiment_dir
        / "ground_truth"
        / "ground_truth_report.json"
    ])

    # --------------------------------------------------------
    # AUTODESK
    # --------------------------------------------------------

    autodesk_report = first_existing([

        experiment_dir
        / "autodesk_output"
        / "autodesk_report.json"
    ])

    autodesk_metrics = first_existing([

        experiment_dir
        / "autodesk_metrics"
        / "autodesk_metrics_report.json",

        experiment_dir
        / "autodesk_output"
        / "autodesk_metrics_report.json"
    ])

    autodesk_step = find_latest_file(

        experiment_dir
        / "autodesk_output",

        [
            "*.step",
            "*.stp"
        ]
    )

    # --------------------------------------------------------
    # BENCHMARK
    # --------------------------------------------------------

    benchmark_dir = (
        experiment_dir
        / "geometry_benchmark"
    )

    our_benchmark = first_existing([

        benchmark_dir
        / "geometric_mapping_geometry_benchmark.json",

        benchmark_dir
        / "our_geometry_benchmark.json"
    ])

    autodesk_benchmark = first_existing([

        benchmark_dir
        / "autodesk_geometry_benchmark.json"
    ])

    return {

        "edit_request":
            edit_request,

        "our_metrics":
            our_metrics,

        "our_edit":
            our_edit,

        "our_module_edit":
            our_module_edit,

        "our_step":
            our_step,

        "ground_truth":
            ground_truth,

        "autodesk_report":
            autodesk_report,

        "autodesk_metrics":
            autodesk_metrics,

        "autodesk_step":
            autodesk_step,

        "our_benchmark":
            our_benchmark,

        "autodesk_benchmark":
            autodesk_benchmark
    }


# ============================================================
# METADATA
# ============================================================

def extract_metadata(
    experiment_id,
    paths
):

    edit_request = load_json(
        paths.get("edit_request")
    )

    our_metrics = load_json(
        paths.get("our_metrics")
    )

    autodesk_report = load_json(
        paths.get("autodesk_report")
    )

    request_id = first_not_none(

        recursive_find(
            edit_request,
            "request_id"
        ),

        recursive_find(
            our_metrics,
            "request_id"
        ),

        recursive_find(
            autodesk_report,
            "request_id"
        )
    )

    operation = first_not_none(

        recursive_find(
            edit_request,
            "operation"
        ),

        recursive_find(
            our_metrics,
            "operation"
        )
    )

    if operation is not None:

        operation = str(
            operation
        ).upper()

    task = first_not_none(

        recursive_find(
            edit_request,
            "instruction"
        ),

        recursive_find(
            edit_request,
            "task"
        ),

        recursive_find(
            edit_request,
            "request_text"
        ),

        recursive_find(
            our_metrics,
            "instruction"
        ),

        recursive_find(
            autodesk_report,
            "instruction"
        )
    )

    return {

        "experiment":
            experiment_id,

        "request_id":
            request_id,

        "operation":
            operation,

        "task":
            task
    }


# ============================================================
# GT
# ============================================================

def extract_gt(report):

    if not isinstance(report, dict):

        return (
            None,
            None
        )

    status = first_not_none(

        report.get("status"),

        recursive_find(
            report,
            "gt_status"
        )
    )

    reason = first_not_none(

        report.get("reason"),

        report.get("message")
    )

    return (
        status,
        reason
    )


# ============================================================
# BENCHMARK METRICS
# ============================================================

def extract_benchmark(report):

    output = {

        "benchmark_status": None,

        "voxel_iou": None,

        "volumetric_precision": None,
        "volumetric_recall": None,
        "volumetric_f1": None,

        "difference_precision": None,
        "difference_recall": None,
        "difference_f1": None,

        "added_f1": None,
        "removed_f1": None,

        "chamfer_distance_mm": None,
        "chamfer_similarity": None
    }

    if not isinstance(report, dict):
        return output

    output[
        "benchmark_status"
    ] = report.get(
        "status"
    )

    benchmark = report.get(
        "benchmark"
    )

    # --------------------------------------------------------
    # STRUCTURED FORMAT
    # --------------------------------------------------------

    if isinstance(
        benchmark,
        dict
    ):

        volumetric = benchmark.get(
            "volumetric",
            {}
        )

        difference = benchmark.get(
            "difference",
            {}
        )

        added = benchmark.get(
            "added",
            {}
        )

        removed = benchmark.get(
            "removed",
            {}
        )

        surface = benchmark.get(
            "surface",
            {}
        )

        output["voxel_iou"] = safe_float(
            volumetric.get("voxel_iou")
        )

        output["volumetric_precision"] = safe_float(
            volumetric.get("precision")
        )

        output["volumetric_recall"] = safe_float(
            volumetric.get("recall")
        )

        output["volumetric_f1"] = safe_float(
            volumetric.get("f1")
        )

        output["difference_precision"] = safe_float(
            difference.get("precision")
        )

        output["difference_recall"] = safe_float(
            difference.get("recall")
        )

        output["difference_f1"] = safe_float(
            difference.get("f1")
        )

        output["added_f1"] = safe_float(
            added.get("f1")
        )

        output["removed_f1"] = safe_float(
            removed.get("f1")
        )

        output["chamfer_distance_mm"] = safe_float(
            surface.get(
                "chamfer_distance_mm"
            )
        )

        output["chamfer_similarity"] = safe_float(
            surface.get(
                "chamfer_similarity"
            )
        )

        return output

    # --------------------------------------------------------
    # FLAT FALLBACK
    # --------------------------------------------------------

    output["voxel_iou"] = safe_float(
        recursive_find(
            report,
            "voxel_iou"
        )
    )

    output["volumetric_precision"] = safe_float(
        recursive_find(
            report,
            "volumetric_precision"
        )
    )

    output["volumetric_recall"] = safe_float(
        recursive_find(
            report,
            "volumetric_recall"
        )
    )

    output["volumetric_f1"] = safe_float(
        recursive_find(
            report,
            "volumetric_f1"
        )
    )

    output["difference_precision"] = safe_float(
        recursive_find(
            report,
            "difference_precision"
        )
    )

    output["difference_recall"] = safe_float(
        recursive_find(
            report,
            "difference_recall"
        )
    )

    output["difference_f1"] = safe_float(
        recursive_find(
            report,
            "difference_f1"
        )
    )

    output["added_f1"] = safe_float(
        recursive_find(
            report,
            "added_f1"
        )
    )

    output["removed_f1"] = safe_float(
        recursive_find(
            report,
            "removed_f1"
        )
    )

    output["chamfer_distance_mm"] = safe_float(
        recursive_find(
            report,
            {
                "chamfer_distance_mm",
                "chamfer_distance"
            }
        )
    )

    output["chamfer_similarity"] = safe_float(
        recursive_find(
            report,
            "chamfer_similarity"
        )
    )

    return output


# ============================================================
# AUTODESK SKIP DETECTION
# ============================================================

def detect_autodesk_skip(report):

    if not isinstance(report, dict):
        return False

    stdout_path = report.get(
        "stdout_log"
    )

    if not stdout_path:
        return False

    try:

        stdout_path = Path(
            stdout_path
        )

        if not stdout_path.exists():
            return False

        text = stdout_path.read_text(
            encoding="utf-8",
            errors="ignore"
        )

        if (
            "Skipping existing request"
            in text
        ):
            return True

        if (
            "Already exists"
            in text
        ):
            return True

    except Exception:
        pass

    return False


# ============================================================
# OUR METHOD
# ============================================================

# ============================================================
# OUR METHOD
# ============================================================

def extract_our_method(
    experiment_id,
    metadata,
    paths
):

    metrics = load_json(
        paths.get("our_metrics")
    )

    edit = load_json(
        paths.get("our_edit")
    )

    module_edit = load_json(
        paths.get("our_module_edit")
    )

    gt_report = load_json(
        paths.get("ground_truth")
    )

    benchmark_report = load_json(
        paths.get("our_benchmark")
    )

    benchmark = extract_benchmark(
        benchmark_report
    )

    gt_status, gt_reason = extract_gt(
        gt_report
    )

    # ========================================================
    # RUNTIME
    # ========================================================
    #
    # REAL metrics_report.json structure:
    #
    # "cad_edit": {
    #     "runtime_s": ...
    # }
    #
    # ========================================================

    runtime = safe_float(
        first_not_none(

            nested_get(
                metrics,
                "cad_edit",
                "runtime_s"
            ),

            nested_get(
                metrics,
                "cad_edit",
                "runtime_sec"
            ),

            nested_get(
                module_edit,
                "runtime_s"
            ),

            nested_get(
                module_edit,
                "runtime_sec"
            ),

            nested_get(
                edit,
                "runtime_s"
            ),

            nested_get(
                edit,
                "runtime_sec"
            )
        )
    )

    # ========================================================
    # OPERATION STATUS
    # ========================================================

    operation_status = first_not_none(

        nested_get(
            metrics,
            "operation_status"
        ),

        nested_get(
            metrics,
            "operation_details",
            "operation_status"
        ),

        nested_get(
            edit,
            "operation_status"
        )
    )

    # ========================================================
    # STEP SUCCESS
    # ========================================================

    step_success = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "cad_edit",
                "step_success"
            ),

            nested_get(
                edit,
                "step_success"
            ),

            (
                True
                if paths.get("our_step")
                else None
            )
        )
    )

    # ========================================================
    # VALID B-REP
    # ========================================================

    valid_brep = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "cad_edit",
                "valid_brep"
            ),

            nested_get(
                metrics,
                "geometry",
                "result",
                "valid"
            ),

            nested_get(
                metrics,
                "success",
                "valid_brep"
            )
        )
    )

    # ========================================================
    # RE-IMPORT
    # ========================================================

    reimport_valid = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "cad_edit",
                "reimport_valid"
            ),

            nested_get(
                metrics,
                "cad_edit",
                "export_reimport_success"
            ),

            nested_get(
                metrics,
                "success",
                "export_reimport_success"
            )
        )
    )

    # ========================================================
    # SOLID COUNT
    # ========================================================

    input_solids = nested_get(
        metrics,
        "geometry",
        "start",
        "solids"
    )

    output_solids = nested_get(
        metrics,
        "geometry",
        "result",
        "solids"
    )

    solid_preserved = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "success",
                "solid_count_preserved"
            ),

            (
                input_solids == output_solids
                if
                input_solids is not None
                and
                output_solids is not None
                else
                None
            )
        )
    )

    # ========================================================
    # FACE COUNT
    # ========================================================

    input_faces = nested_get(
        metrics,
        "geometry",
        "start",
        "faces"
    )

    output_faces = nested_get(
        metrics,
        "geometry",
        "result",
        "faces"
    )

    # ========================================================
    # EDGE COUNT
    # ========================================================

    input_edges = nested_get(
        metrics,
        "geometry",
        "start",
        "edges"
    )

    output_edges = nested_get(
        metrics,
        "geometry",
        "result",
        "edges"
    )

    # ========================================================
    # FINAL MODEL VOLUME CHANGE
    # ========================================================
    #
    # CRITICAL:
    #
    # Use geometry.change.volume_delta_mm3.
    #
    # DO NOT recursively search for volume_change_mm3,
    # because operation_details.chamfer.contours contains
    # intermediate contour-level volume changes.
    #
    # ========================================================

    volume_change = safe_float(
        nested_get(
            metrics,
            "geometry",
            "change",
            "volume_delta_mm3"
        )
    )

    relative_volume_change = safe_float(
        nested_get(
            metrics,
            "geometry",
            "change",
            "relative_volume_change"
        )
    )

    # ========================================================
    # ADAPTIVE OPERATION DETAILS
    # ========================================================

    operation = str(
        first_not_none(
            metadata.get("operation"),
            metrics.get("operation")
            if isinstance(metrics, dict)
            else None
        )
        or ""
    ).upper()

    requested_distance = None

    minimum_applied = None
    maximum_applied = None
    mean_applied = None

    minimum_retention = None
    mean_retention = None

    contour_count = None
    skipped_edges = None

    # --------------------------------------------------------
    # CHAMFER
    # --------------------------------------------------------

    if operation == "CHAMFER":

        chamfer = nested_get(
            metrics,
            "operation_details",
            "chamfer"
        )

        if isinstance(chamfer, dict):

            requested_distance = safe_float(
                chamfer.get(
                    "requested_distance_mm"
                )
            )

            minimum_applied = safe_float(
                chamfer.get(
                    "minimum_applied_distance_mm"
                )
            )

            maximum_applied = safe_float(
                chamfer.get(
                    "maximum_applied_distance_mm"
                )
            )

            mean_applied = safe_float(
                chamfer.get(
                    "mean_applied_distance_mm"
                )
            )

            minimum_retention = safe_float(
                chamfer.get(
                    "minimum_distance_retention"
                )
            )

            mean_retention = safe_float(
                chamfer.get(
                    "mean_distance_retention"
                )
            )

            contour_count = safe_int(
                chamfer.get(
                    "contours_total"
                )
            )

            skipped = chamfer.get(
                "skipped_edges"
            )

            if isinstance(skipped, list):

                skipped_edges = len(
                    skipped
                )

            else:

                skipped_edges = safe_int(
                    skipped
                )

    # --------------------------------------------------------
    # FILLET
    # --------------------------------------------------------
    #
    # Supports future metrics_report structure:
    #
    # operation_details.fillet
    #
    # --------------------------------------------------------

    elif operation == "FILLET":

        fillet = nested_get(
            metrics,
            "operation_details",
            "fillet"
        )

        if isinstance(fillet, dict):

            requested_distance = safe_float(
                first_not_none(
                    fillet.get(
                        "requested_radius_mm"
                    ),
                    fillet.get(
                        "requested_distance_mm"
                    )
                )
            )

            minimum_applied = safe_float(
                first_not_none(
                    fillet.get(
                        "minimum_applied_radius_mm"
                    ),
                    fillet.get(
                        "minimum_applied_distance_mm"
                    )
                )
            )

            maximum_applied = safe_float(
                first_not_none(
                    fillet.get(
                        "maximum_applied_radius_mm"
                    ),
                    fillet.get(
                        "maximum_applied_distance_mm"
                    )
                )
            )

            mean_applied = safe_float(
                first_not_none(
                    fillet.get(
                        "mean_applied_radius_mm"
                    ),
                    fillet.get(
                        "mean_applied_distance_mm"
                    )
                )
            )

            minimum_retention = safe_float(
                first_not_none(
                    fillet.get(
                        "minimum_radius_retention"
                    ),
                    fillet.get(
                        "minimum_distance_retention"
                    )
                )
            )

            mean_retention = safe_float(
                first_not_none(
                    fillet.get(
                        "mean_radius_retention"
                    ),
                    fillet.get(
                        "mean_distance_retention"
                    )
                )
            )

            contour_count = safe_int(
                first_not_none(
                    fillet.get(
                        "contours_total"
                    ),
                    fillet.get(
                        "groups_total"
                    )
                )
            )

            skipped = fillet.get(
                "skipped_edges"
            )

            if isinstance(skipped, list):

                skipped_edges = len(
                    skipped
                )

            else:

                skipped_edges = safe_int(
                    skipped
                )

    # ========================================================
    # RETURN
    # ========================================================

    return {

        "experiment":
            experiment_id,

        "request_id":
            metadata.get(
                "request_id"
            ),

        "operation":
            metadata.get(
                "operation"
            ),

        "task":
            metadata.get(
                "task"
            ),

        "method":
            "GEOMETRIC_MAPPING",

        "fresh_inference":
            None,

        "operation_status":
            operation_status,

        "step_success":
            step_success,

        "valid_brep":
            valid_brep,

        "reimport_valid":
            reimport_valid,

        "solid_count_preserved":
            solid_preserved,

        "input_solids":
            input_solids,

        "output_solids":
            output_solids,

        "input_faces":
            input_faces,

        "output_faces":
            output_faces,

        "input_edges":
            input_edges,

        "output_edges":
            output_edges,

        "runtime_sec":
            runtime,

        # ----------------------------------------------------
        # Deterministic execution:
        # no VLM inference loop required.
        # ----------------------------------------------------

        "vlm_responses":
            0,

        "visual_renders":
            0,

        "input_tokens":
            0,

        "output_tokens":
            0,

        "total_tokens":
            0,

        "cost_usd":
            0.0,

        "volume_change_mm3":
            volume_change,

        "relative_volume_change":
            relative_volume_change,

        "gt_status":
            gt_status,

        "gt_reason":
            gt_reason,

        **benchmark,

        "requested_distance_mm":
            requested_distance,

        "minimum_applied_distance_mm":
            minimum_applied,

        "maximum_applied_distance_mm":
            maximum_applied,

        "mean_applied_distance_mm":
            mean_applied,

        "minimum_distance_retention":
            minimum_retention,

        "mean_distance_retention":
            mean_retention,

        "contour_count":
            contour_count,

        "skipped_edges":
            skipped_edges,

        "result_step":
            (
                str(
                    paths["our_step"]
                )
                if paths.get(
                    "our_step"
                )
                else None
            )
    }

# ============================================================
# AUTODESK METHOD
# ============================================================

def extract_autodesk_method(
    experiment_id,
    metadata,
    paths
):

    report = load_json(
        paths.get(
            "autodesk_report"
        )
    )

    metrics = load_json(
        paths.get(
            "autodesk_metrics"
        )
    )

    gt_report = load_json(
        paths.get(
            "ground_truth"
        )
    )

    benchmark_report = load_json(
        paths.get(
            "autodesk_benchmark"
        )
    )

    benchmark = extract_benchmark(
        benchmark_report
    )

    gt_status, gt_reason = extract_gt(
        gt_report
    )

    skipped_existing = detect_autodesk_skip(
        report
    )

    # ========================================================
    # RUNTIME
    # ========================================================

    runtime = safe_float(
        first_not_none(

            nested_get(
                metrics,
                "runtime",
                "sec"
            ),

            nested_get(
                metrics,
                "runtime_sec"
            ),

            nested_get(
                report,
                "runtime_sec"
            )
        )
    )

    # A cached launcher runtime is not inference runtime.

    if skipped_existing:
        runtime = None

    # ========================================================
    # STATUS
    # ========================================================

    operation_status = first_not_none(

        nested_get(
            metrics,
            "status"
        ),

        nested_get(
            report,
            "status"
        )
    )

    if skipped_existing:
        operation_status = "SKIPPED_EXISTING"

    # ========================================================
    # STEP
    # ========================================================

    step_success = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "step",
                "success"
            ),

            nested_get(
                metrics,
                "step_success"
            ),

            nested_get(
                report,
                "step_success"
            ),

            (
                True
                if paths.get(
                    "autodesk_step"
                )
                else None
            )
        )
    )

    valid_brep = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "step",
                "valid_brep"
            ),

            nested_get(
                metrics,
                "valid_brep"
            ),

            recursive_find(
                metrics,
                "result_valid"
            )
        )
    )

    reimport_valid = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "step",
                "reimport_valid"
            ),

            nested_get(
                metrics,
                "reimport_valid"
            ),

            valid_brep
        )
    )

    solid_preserved = safe_bool(
        first_not_none(

            nested_get(
                metrics,
                "step",
                "solid_count_preserved"
            ),

            nested_get(
                metrics,
                "solid_count_preserved"
            )
        )
    )

    # ========================================================
    # GEOMETRY
    # ========================================================

    input_solids = first_not_none(

        nested_get(
            metrics,
            "geometry",
            "start",
            "solids"
        ),

        recursive_find(
            metrics,
            "input_solids"
        )
    )

    output_solids = first_not_none(

        nested_get(
            metrics,
            "geometry",
            "result",
            "solids"
        ),

        recursive_find(
            metrics,
            "output_solids"
        )
    )

    input_faces = first_not_none(

        nested_get(
            metrics,
            "geometry",
            "start",
            "faces"
        ),

        recursive_find(
            metrics,
            "input_faces"
        )
    )

    output_faces = first_not_none(

        nested_get(
            metrics,
            "geometry",
            "result",
            "faces"
        ),

        recursive_find(
            metrics,
            "output_faces"
        )
    )

    input_edges = first_not_none(

        nested_get(
            metrics,
            "geometry",
            "start",
            "edges"
        ),

        recursive_find(
            metrics,
            "input_edges"
        )
    )

    output_edges = first_not_none(

        nested_get(
            metrics,
            "geometry",
            "result",
            "edges"
        ),

        recursive_find(
            metrics,
            "output_edges"
        )
    )

    if (
        solid_preserved is None
        and
        input_solids is not None
        and
        output_solids is not None
    ):

        solid_preserved = (
            input_solids
            ==
            output_solids
        )

    # ========================================================
    # VOLUME
    # ========================================================

    volume_change = safe_float(
        first_not_none(

            nested_get(
                metrics,
                "geometry",
                "change",
                "volume_delta_mm3"
            ),

            nested_get(
                metrics,
                "geometry",
                "change",
                "volume_change_mm3"
            )
        )
    )

    relative_volume_change = safe_float(
        nested_get(
            metrics,
            "geometry",
            "change",
            "relative_volume_change"
        )
    )

    # ========================================================
    # AGENT
    # ========================================================

    vlm_responses = safe_int(
        first_not_none(

            nested_get(
                metrics,
                "agent",
                "vlm_responses"
            ),

            recursive_find(
                metrics,
                "vlm_responses"
            )
        )
    )

    visual_renders = safe_int(
        first_not_none(

            nested_get(
                metrics,
                "agent",
                "visual_feedback_renders"
            ),

            nested_get(
                metrics,
                "agent",
                "visual_renders"
            )
        )
    )

    input_tokens = safe_int(
        nested_get(
            metrics,
            "agent",
            "input_tokens"
        )
    )

    output_tokens = safe_int(
        nested_get(
            metrics,
            "agent",
            "output_tokens"
        )
    )

    total_tokens = safe_int(
        nested_get(
            metrics,
            "agent",
            "total_tokens"
        )
    )

    cost_usd = safe_float(
        first_not_none(

            nested_get(
                metrics,
                "agent",
                "cost_estimate_usd"
            ),

            nested_get(
                metrics,
                "agent",
                "cost_usd"
            )
        )
    )

    return {

        "experiment":
            experiment_id,

        "request_id":
            metadata.get(
                "request_id"
            ),

        "operation":
            metadata.get(
                "operation"
            ),

        "task":
            metadata.get(
                "task"
            ),

        "method":
            "AUTODESK",

        "fresh_inference":
            (
                False
                if skipped_existing
                else True
            ),

        "operation_status":
            operation_status,

        "step_success":
            step_success,

        "valid_brep":
            valid_brep,

        "reimport_valid":
            reimport_valid,

        "solid_count_preserved":
            solid_preserved,

        "input_solids":
            input_solids,

        "output_solids":
            output_solids,

        "input_faces":
            input_faces,

        "output_faces":
            output_faces,

        "input_edges":
            input_edges,

        "output_edges":
            output_edges,

        "runtime_sec":
            runtime,

        "vlm_responses":
            vlm_responses,

        "visual_renders":
            visual_renders,

        "input_tokens":
            input_tokens,

        "output_tokens":
            output_tokens,

        "total_tokens":
            total_tokens,

        "cost_usd":
            cost_usd,

        "volume_change_mm3":
            volume_change,

        "relative_volume_change":
            relative_volume_change,

        "gt_status":
            gt_status,

        "gt_reason":
            gt_reason,

        **benchmark,

        "requested_distance_mm":
            None,

        "minimum_applied_distance_mm":
            None,

        "maximum_applied_distance_mm":
            None,

        "mean_applied_distance_mm":
            None,

        "minimum_distance_retention":
            None,

        "mean_distance_retention":
            None,

        "contour_count":
            None,

        "skipped_edges":
            None,

        "result_step":
            (
                str(
                    paths[
                        "autodesk_step"
                    ]
                )
                if paths.get(
                    "autodesk_step"
                )
                else None
            )
    }


# ============================================================
# COLLECT
# ============================================================

def collect_results(
    experiments
):

    rows = []
    availability = []
    metadata_rows = []

    for (
        experiment_number,
        experiment_id,
        experiment_dir
    ) in experiments:

        header(
            f"READ EXPERIMENT {experiment_id}"
        )

        paths = discover_paths(
            experiment_id,
            experiment_dir
        )

        metadata = extract_metadata(
            experiment_id,
            paths
        )

        metadata_rows.append(
            metadata
        )

        print(
            "Request ID:",
            metadata.get(
                "request_id"
            )
        )

        print(
            "Operation:",
            metadata.get(
                "operation"
            )
        )

        for name, path in paths.items():

            exists = (
                path is not None
                and
                Path(path).exists()
            )

            print(
                f"{name:24s}: {exists}"
            )

            availability.append({

                "experiment":
                    experiment_id,

                "report":
                    name,

                "exists":
                    exists,

                "path":
                    (
                        str(path)
                        if path is not None
                        else None
                    )
            })

        our_row = extract_our_method(
            experiment_id,
            metadata,
            paths
        )

        rows.append(
            our_row
        )

        autodesk_row = extract_autodesk_method(
            experiment_id,
            metadata,
            paths
        )

        rows.append(
            autodesk_row
        )

    return (
        rows,
        availability,
        metadata_rows
    )


# ============================================================
# SUMMARY HELPERS
# ============================================================

def numeric_values(
    rows,
    field
):

    values = []

    for row in rows:

        value = safe_float(
            row.get(
                field
            )
        )

        if value is not None:

            values.append(
                value
            )

    return values


def count_true(
    rows,
    field
):

    return sum(
        row.get(field) is True
        for row in rows
    )


def count_false(
    rows,
    field
):

    return sum(
        row.get(field) is False
        for row in rows
    )


# ============================================================
# SPEEDUP
# ============================================================

def calculate_speedups(rows):

    experiment_ids = sorted(

        set(
            row["experiment"]
            for row in rows
        ),

        key=lambda name:
            int(
                re.search(
                    r"\d+",
                    name
                ).group()
            )
    )

    result = {}

    for experiment_id in experiment_ids:

        our = next(

            (
                row
                for row in rows

                if
                row["experiment"]
                ==
                experiment_id

                and
                row["method"]
                ==
                "GEOMETRIC_MAPPING"
            ),

            None
        )

        autodesk = next(

            (
                row
                for row in rows

                if
                row["experiment"]
                ==
                experiment_id

                and
                row["method"]
                ==
                "AUTODESK"
            ),

            None
        )

        if our is None or autodesk is None:

            result[
                experiment_id
            ] = None

            continue

        if (
            autodesk.get(
                "fresh_inference"
            )
            is False
        ):

            result[
                experiment_id
            ] = None

            continue

        our_runtime = safe_float(
            our.get(
                "runtime_sec"
            )
        )

        autodesk_runtime = safe_float(
            autodesk.get(
                "runtime_sec"
            )
        )

        if (
            our_runtime is None
            or
            autodesk_runtime is None
            or
            our_runtime <= 0
        ):

            result[
                experiment_id
            ] = None

            continue

        result[
            experiment_id
        ] = (
            autodesk_runtime
            /
            our_runtime
        )

    return result


# ============================================================
# METHOD SUMMARY
# ============================================================

def build_method_summary(rows):

    methods = sorted(
        set(
            row["method"]
            for row in rows
        )
    )

    summary = []

    for method in methods:

        method_rows = [

            row
            for row in rows
            if row["method"] == method
        ]

        runtime_rows = method_rows

        if method == "AUTODESK":

            runtime_rows = [

                row
                for row in method_rows

                if
                row.get(
                    "fresh_inference"
                )
                is not False
            ]

        runtimes = numeric_values(
            runtime_rows,
            "runtime_sec"
        )

        vlm = numeric_values(
            method_rows,
            "vlm_responses"
        )

        renders = numeric_values(
            method_rows,
            "visual_renders"
        )

        tokens = numeric_values(
            method_rows,
            "total_tokens"
        )

        costs = numeric_values(
            method_rows,
            "cost_usd"
        )

        voxel_iou = numeric_values(
            method_rows,
            "voxel_iou"
        )

        volumetric_f1 = numeric_values(
            method_rows,
            "volumetric_f1"
        )

        difference_f1 = numeric_values(
            method_rows,
            "difference_f1"
        )

        chamfer = numeric_values(
            method_rows,
            "chamfer_distance_mm"
        )

        total_runtime = (
            sum(runtimes)
            if runtimes
            else None
        )

        summary.append({

            "method":
                method,

            "experiments":
                len(
                    method_rows
                ),

            "step_success":
                count_true(
                    method_rows,
                    "step_success"
                ),

            "valid_brep":
                count_true(
                    method_rows,
                    "valid_brep"
                ),

            "solid_preserved":
                count_true(
                    method_rows,
                    "solid_count_preserved"
                ),

            "solid_not_preserved":
                count_false(
                    method_rows,
                    "solid_count_preserved"
                ),

            "fresh_runtime_count":
                len(
                    runtimes
                ),

            "mean_runtime_sec":
                (
                    mean(runtimes)
                    if runtimes
                    else None
                ),

            "median_runtime_sec":
                (
                    median(runtimes)
                    if runtimes
                    else None
                ),

            "min_runtime_sec":
                (
                    min(runtimes)
                    if runtimes
                    else None
                ),

            "max_runtime_sec":
                (
                    max(runtimes)
                    if runtimes
                    else None
                ),

            "total_runtime_sec":
                total_runtime,

            "total_vlm_responses":
                (
                    sum(vlm)
                    if vlm
                    else None
                ),

            "total_visual_renders":
                (
                    sum(renders)
                    if renders
                    else None
                ),

            "total_tokens":
                (
                    sum(tokens)
                    if tokens
                    else None
                ),

            "total_cost_usd":
                (
                    sum(costs)
                    if costs
                    else None
                ),

            "gt_metric_count":
                len(
                    voxel_iou
                ),

            "mean_voxel_iou":
                (
                    mean(voxel_iou)
                    if voxel_iou
                    else None
                ),

            "mean_volumetric_f1":
                (
                    mean(volumetric_f1)
                    if volumetric_f1
                    else None
                ),

            "mean_difference_f1":
                (
                    mean(difference_f1)
                    if difference_f1
                    else None
                ),

            "mean_chamfer_distance_mm":
                (
                    mean(chamfer)
                    if chamfer
                    else None
                )
        })

    return summary


# ============================================================
# EXCEL STYLES
# ============================================================

HEADER_FILL = PatternFill(
    "solid",
    fgColor="1F4E78"
)

GOOD_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9"
)

BAD_FILL = PatternFill(
    "solid",
    fgColor="FCE4D6"
)

NA_FILL = PatternFill(
    "solid",
    fgColor="E7E6E6"
)

THIN_BORDER = Border(

    left=Side(
        style="thin",
        color="D9D9D9"
    ),

    right=Side(
        style="thin",
        color="D9D9D9"
    ),

    top=Side(
        style="thin",
        color="D9D9D9"
    ),

    bottom=Side(
        style="thin",
        color="D9D9D9"
    )
)


def format_header(ws):

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = HEADER_FILL

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

    ws.row_dimensions[
        1
    ].height = 40


def format_cells(ws):

    for row in ws.iter_rows():

        for cell in row:

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


def autofit_columns(
    ws,
    max_width=55
):

    for column_cells in ws.columns:

        column_letter = get_column_letter(
            column_cells[
                0
            ].column
        )

        max_length = 0

        for cell in column_cells:

            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(
                    str(
                        cell.value
                    )
                )
            )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            max_width
        )


def format_boolean_columns(
    ws,
    column_names
):

    columns = {

        cell.value:
            cell.column

        for cell in ws[1]
    }

    for column_name in column_names:

        column = columns.get(
            column_name
        )

        if column is None:
            continue

        for row_number in range(
            2,
            ws.max_row + 1
        ):

            cell = ws.cell(
                row=row_number,
                column=column
            )

            if cell.value == "YES":

                cell.fill = GOOD_FILL

            elif cell.value == "NO":

                cell.fill = BAD_FILL

            elif cell.value == "N/A":

                cell.fill = NA_FILL


# ============================================================
# SHEET 1
# ============================================================

def create_experiment_sheet(
    workbook,
    rows
):

    ws = workbook.active

    ws.title = (
        "Experiment Comparison"
    )

    speedups = calculate_speedups(
        rows
    )

    headers = [

        "Experiment",
        "Request ID",
        "Operation",
        "Task",

        "Method",

        "Fresh Inference",

        "Operation Status",

        "STEP Success",
        "Valid B-Rep",
        "Re-import Valid",
        "Solid Count Preserved",

        "Input Solids",
        "Output Solids",

        "Input Faces",
        "Output Faces",

        "Input Edges",
        "Output Edges",

        "Runtime [s]",
        "Runtime [min]",

        "Autodesk / OUR Speedup",

        "VLM Responses",
        "Visual Renders",

        "Input Tokens",
        "Output Tokens",
        "Total Tokens",

        "Estimated Cost [USD]",

        "Volume Change [mm³]",
        "Relative Volume Change",

        "GT Status",

        "Voxel IoU",
        "Volumetric F1",
        "Difference F1",

        "Added F1",
        "Removed F1",

        "Chamfer Distance [mm]",

        "Result STEP"
    ]

    ws.append(
        headers
    )

    for row in rows:

        runtime = safe_float(
            row.get(
                "runtime_sec"
            )
        )

        if row["method"] == "AUTODESK":

            fresh = excel_bool(
                row.get(
                    "fresh_inference"
                )
            )

        else:

            fresh = "N/A"

        ws.append([

            row.get(
                "experiment"
            ),

            row.get(
                "request_id"
            ),

            row.get(
                "operation"
            ),

            row.get(
                "task"
            ),

            row.get(
                "method"
            ),

            fresh,

            row.get(
                "operation_status"
            ),

            excel_bool(
                row.get(
                    "step_success"
                )
            ),

            excel_bool(
                row.get(
                    "valid_brep"
                )
            ),

            excel_bool(
                row.get(
                    "reimport_valid"
                )
            ),

            excel_bool(
                row.get(
                    "solid_count_preserved"
                )
            ),

            row.get(
                "input_solids"
            ),

            row.get(
                "output_solids"
            ),

            row.get(
                "input_faces"
            ),

            row.get(
                "output_faces"
            ),

            row.get(
                "input_edges"
            ),

            row.get(
                "output_edges"
            ),

            runtime,

            (
                runtime / 60.0
                if runtime is not None
                else None
            ),

            speedups.get(
                row.get(
                    "experiment"
                )
            ),

            row.get(
                "vlm_responses"
            ),

            row.get(
                "visual_renders"
            ),

            row.get(
                "input_tokens"
            ),

            row.get(
                "output_tokens"
            ),

            row.get(
                "total_tokens"
            ),

            row.get(
                "cost_usd"
            ),

            row.get(
                "volume_change_mm3"
            ),

            row.get(
                "relative_volume_change"
            ),

            row.get(
                "gt_status"
            ),

            row.get(
                "voxel_iou"
            ),

            row.get(
                "volumetric_f1"
            ),

            row.get(
                "difference_f1"
            ),

            row.get(
                "added_f1"
            ),

            row.get(
                "removed_f1"
            ),

            row.get(
                "chamfer_distance_mm"
            ),

            row.get(
                "result_step"
            )
        ])

    format_header(
        ws
    )

    format_cells(
        ws
    )

    format_boolean_columns(
        ws,
        [
            "Fresh Inference",
            "STEP Success",
            "Valid B-Rep",
            "Re-import Valid",
            "Solid Count Preserved"
        ]
    )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    autofit_columns(
        ws,
        60
    )

    return ws


# ============================================================
# SHEET 2
# ============================================================

def create_summary_sheet(
    workbook,
    rows
):

    ws = workbook.create_sheet(
        "Method Summary"
    )

    summary = build_method_summary(
        rows
    )

    headers = [

        "Method",

        "Experiments",

        "STEP Success Count",

        "Valid B-Rep Count",

        "Solid Preserved Count",

        "Solid NOT Preserved Count",

        "Fresh Runtime Count",

        "Mean Runtime [s]",

        "Median Runtime [s]",

        "Min Runtime [s]",

        "Max Runtime [s]",

        "Total Runtime [s]",

        "Total Runtime [min]",

        "Total VLM Responses",

        "Total Visual Renders",

        "Total Tokens",

        "Total Cost [USD]",

        "GT Metric Count",

        "Mean Voxel IoU",

        "Mean Volumetric F1",

        "Mean Difference F1",

        "Mean Chamfer Distance [mm]"
    ]

    ws.append(
        headers
    )

    for row in summary:

        total_runtime = row.get(
            "total_runtime_sec"
        )

        ws.append([

            row.get(
                "method"
            ),

            row.get(
                "experiments"
            ),

            row.get(
                "step_success"
            ),

            row.get(
                "valid_brep"
            ),

            row.get(
                "solid_preserved"
            ),

            row.get(
                "solid_not_preserved"
            ),

            row.get(
                "fresh_runtime_count"
            ),

            row.get(
                "mean_runtime_sec"
            ),

            row.get(
                "median_runtime_sec"
            ),

            row.get(
                "min_runtime_sec"
            ),

            row.get(
                "max_runtime_sec"
            ),

            total_runtime,

            (
                total_runtime / 60.0
                if total_runtime is not None
                else None
            ),

            row.get(
                "total_vlm_responses"
            ),

            row.get(
                "total_visual_renders"
            ),

            row.get(
                "total_tokens"
            ),

            row.get(
                "total_cost_usd"
            ),

            row.get(
                "gt_metric_count"
            ),

            row.get(
                "mean_voxel_iou"
            ),

            row.get(
                "mean_volumetric_f1"
            ),

            row.get(
                "mean_difference_f1"
            ),

            row.get(
                "mean_chamfer_distance_mm"
            )
        ])

    format_header(
        ws
    )

    format_cells(
        ws
    )

    ws.freeze_panes = "A2"

    autofit_columns(
        ws
    )

    return ws


# ============================================================
# SHEET 3
# ============================================================

def create_quality_sheet(
    workbook,
    rows
):

    ws = workbook.create_sheet(
        "Quality Metrics"
    )

    headers = [

        "Experiment",
        "Operation",
        "Method",

        "GT Status",
        "GT Reason",

        "Benchmark Status",

        "Voxel IoU",

        "Volumetric Precision",
        "Volumetric Recall",
        "Volumetric F1",

        "Difference Precision",
        "Difference Recall",
        "Difference F1",

        "Added F1",
        "Removed F1",

        "Chamfer Distance [mm]",
        "Chamfer Similarity"
    ]

    ws.append(
        headers
    )

    for row in rows:

        ws.append([

            row.get(
                "experiment"
            ),

            row.get(
                "operation"
            ),

            row.get(
                "method"
            ),

            row.get(
                "gt_status"
            ),

            row.get(
                "gt_reason"
            ),

            row.get(
                "benchmark_status"
            ),

            row.get(
                "voxel_iou"
            ),

            row.get(
                "volumetric_precision"
            ),

            row.get(
                "volumetric_recall"
            ),

            row.get(
                "volumetric_f1"
            ),

            row.get(
                "difference_precision"
            ),

            row.get(
                "difference_recall"
            ),

            row.get(
                "difference_f1"
            ),

            row.get(
                "added_f1"
            ),

            row.get(
                "removed_f1"
            ),

            row.get(
                "chamfer_distance_mm"
            ),

            row.get(
                "chamfer_similarity"
            )
        ])

    format_header(
        ws
    )

    format_cells(
        ws
    )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    autofit_columns(
        ws,
        70
    )

    return ws


# ============================================================
# SHEET 4
# ============================================================

def create_adaptive_sheet(
    workbook,
    rows
):

    ws = workbook.create_sheet(
        "Adaptive Details"
    )

    headers = [

        "Experiment",
        "Operation",
        "Method",

        "Operation Status",

        "Requested Distance [mm]",

        "Minimum Applied [mm]",
        "Maximum Applied [mm]",
        "Mean Applied [mm]",

        "Minimum Retention",
        "Mean Retention",

        "Contour Count",

        "Skipped Edges"
    ]

    ws.append(
        headers
    )

    for row in rows:

        ws.append([

            row.get(
                "experiment"
            ),

            row.get(
                "operation"
            ),

            row.get(
                "method"
            ),

            row.get(
                "operation_status"
            ),

            row.get(
                "requested_distance_mm"
            ),

            row.get(
                "minimum_applied_distance_mm"
            ),

            row.get(
                "maximum_applied_distance_mm"
            ),

            row.get(
                "mean_applied_distance_mm"
            ),

            row.get(
                "minimum_distance_retention"
            ),

            row.get(
                "mean_distance_retention"
            ),

            row.get(
                "contour_count"
            ),

            row.get(
                "skipped_edges"
            )
        ])

    format_header(
        ws
    )

    format_cells(
        ws
    )

    ws.freeze_panes = "A2"

    autofit_columns(
        ws
    )

    return ws


# ============================================================
# SHEET 5
# ============================================================

def create_availability_sheet(
    workbook,
    availability
):

    ws = workbook.create_sheet(
        "Data Availability"
    )

    headers = [

        "Experiment",
        "Report / File",
        "Exists",
        "Path"
    ]

    ws.append(
        headers
    )

    for item in availability:

        ws.append([

            item.get(
                "experiment"
            ),

            item.get(
                "report"
            ),

            excel_bool(
                item.get(
                    "exists"
                )
            ),

            item.get(
                "path"
            )
        ])

    format_header(
        ws
    )

    format_cells(
        ws
    )

    format_boolean_columns(
        ws,
        [
            "Exists"
        ]
    )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    autofit_columns(
        ws,
        100
    )

    return ws


# ============================================================
# SHEET 6
# ============================================================

def create_index_sheet(
    workbook,
    metadata_rows
):

    ws = workbook.create_sheet(
        "Experiment Index"
    )

    headers = [

        "Experiment",
        "Request ID",
        "Operation",
        "Task"
    ]

    ws.append(
        headers
    )

    for row in metadata_rows:

        ws.append([

            row.get(
                "experiment"
            ),

            row.get(
                "request_id"
            ),

            row.get(
                "operation"
            ),

            row.get(
                "task"
            )
        ])

    format_header(
        ws
    )

    format_cells(
        ws
    )

    ws.freeze_panes = "A2"

    autofit_columns(
        ws,
        100
    )

    return ws


# ============================================================
# NUMBER FORMATS
# ============================================================

def apply_number_formats(
    workbook
):

    two_decimal = {

        "Runtime [s]",
        "Runtime [min]",

        "Autodesk / OUR Speedup",

        "Mean Runtime [s]",
        "Median Runtime [s]",
        "Min Runtime [s]",
        "Max Runtime [s]",

        "Total Runtime [s]",
        "Total Runtime [min]",

        "Volume Change [mm³]",

        "Chamfer Distance [mm]",
        "Mean Chamfer Distance [mm]",

        "Requested Distance [mm]",

        "Minimum Applied [mm]",
        "Maximum Applied [mm]",
        "Mean Applied [mm]"
    }

    four_decimal = {

        "Relative Volume Change",

        "Voxel IoU",

        "Volumetric Precision",
        "Volumetric Recall",
        "Volumetric F1",

        "Difference Precision",
        "Difference Recall",
        "Difference F1",

        "Added F1",
        "Removed F1",

        "Chamfer Similarity",

        "Mean Voxel IoU",
        "Mean Volumetric F1",
        "Mean Difference F1",

        "Minimum Retention",
        "Mean Retention"
    }

    money = {

        "Estimated Cost [USD]",
        "Total Cost [USD]"
    }

    for ws in workbook.worksheets:

        headers = {

            cell.value:
                cell.column

            for cell in ws[1]
        }

        for name in two_decimal:

            column = headers.get(
                name
            )

            if column is None:
                continue

            for row in range(
                2,
                ws.max_row + 1
            ):

                ws.cell(
                    row=row,
                    column=column
                ).number_format = (
                    "0.00"
                )

        for name in four_decimal:

            column = headers.get(
                name
            )

            if column is None:
                continue

            for row in range(
                2,
                ws.max_row + 1
            ):

                ws.cell(
                    row=row,
                    column=column
                ).number_format = (
                    "0.0000"
                )

        for name in money:

            column = headers.get(
                name
            )

            if column is None:
                continue

            for row in range(
                2,
                ws.max_row + 1
            ):

                ws.cell(
                    row=row,
                    column=column
                ).number_format = (
                    "$0.0000"
                )


# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel(
    rows,
    availability,
    metadata_rows
):

    COMPARISON_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook = Workbook()

    create_experiment_sheet(
        workbook,
        rows
    )

    create_summary_sheet(
        workbook,
        rows
    )

    create_quality_sheet(
        workbook,
        rows
    )

    create_adaptive_sheet(
        workbook,
        rows
    )

    create_availability_sheet(
        workbook,
        availability
    )

    create_index_sheet(
        workbook,
        metadata_rows
    )

    apply_number_formats(
        workbook
    )

    workbook.save(
        OUTPUT_XLSX
    )

    return OUTPUT_XLSX


# ============================================================
# CONSOLE SUMMARY
# ============================================================

def print_summary(
    experiments,
    rows
):

    header(
        "FINAL EXPERIMENT SUMMARY"
    )

    speedups = calculate_speedups(
        rows
    )

    for (
        experiment_number,
        experiment_id,
        experiment_dir
    ) in experiments:

        print(
            "\n"
            +
            "-" * 100
        )

        print(
            experiment_id
        )

        our = next(

            (
                row
                for row in rows

                if
                row["experiment"]
                ==
                experiment_id

                and
                row["method"]
                ==
                "GEOMETRIC_MAPPING"
            ),

            None
        )

        autodesk = next(

            (
                row
                for row in rows

                if
                row["experiment"]
                ==
                experiment_id

                and
                row["method"]
                ==
                "AUTODESK"
            ),

            None
        )

        if our:

            print(
                "\nOUR METHOD"
            )

            print(
                "Status:",
                our.get(
                    "operation_status"
                )
            )

            print(
                "Runtime:",
                our.get(
                    "runtime_sec"
                )
            )

            print(
                "Final volume change:",
                our.get(
                    "volume_change_mm3"
                )
            )

            print(
                "Valid B-Rep:",
                our.get(
                    "valid_brep"
                )
            )

            print(
                "Solid preserved:",
                our.get(
                    "solid_count_preserved"
                )
            )

        if autodesk:

            print(
                "\nAUTODESK"
            )

            print(
                "Status:",
                autodesk.get(
                    "operation_status"
                )
            )

            print(
                "Fresh inference:",
                autodesk.get(
                    "fresh_inference"
                )
            )

            print(
                "Runtime:",
                autodesk.get(
                    "runtime_sec"
                )
            )

            print(
                "VLM responses:",
                autodesk.get(
                    "vlm_responses"
                )
            )

            print(
                "Visual renders:",
                autodesk.get(
                    "visual_renders"
                )
            )

            print(
                "Valid B-Rep:",
                autodesk.get(
                    "valid_brep"
                )
            )

            print(
                "Solid preserved:",
                autodesk.get(
                    "solid_count_preserved"
                )
            )

        print(
            "\nAutodesk / OUR speedup:",
            speedups.get(
                experiment_id
            )
        )


# ============================================================
# MAIN
# ============================================================

def main():

    header(
        "ASME CIE 2026 — UNIVERSAL METHOD COMPARISON"
    )

    print(
        "Project root:"
    )

    print(
        PROJECT_ROOT
    )

    print(
        "\nExperiments directory:"
    )

    print(
        EXPERIMENTS_DIR
    )

    # --------------------------------------------------------
    # DISCOVER
    # --------------------------------------------------------

    experiments = discover_experiments()

    if not experiments:

        raise RuntimeError(
            "No experiment folders Bxx found."
        )

    header(
        "EXPERIMENT DISCOVERY"
    )

    print(
        "Experiments found:",
        len(
            experiments
        )
    )

    for (
        number,
        experiment_id,
        experiment_dir
    ) in experiments:

        print(
            f"{experiment_id:10s}"
            " -> "
            f"{experiment_dir}"
        )

    # --------------------------------------------------------
    # COLLECT
    # --------------------------------------------------------

    (
        rows,
        availability,
        metadata_rows
    ) = collect_results(
        experiments
    )

    # --------------------------------------------------------
    # CONSOLE
    # --------------------------------------------------------

    print_summary(
        experiments,
        rows
    )

    # --------------------------------------------------------
    # EXCEL
    # --------------------------------------------------------

    output = create_excel(
        rows,
        availability,
        metadata_rows
    )

    # --------------------------------------------------------
    # COMPLETE
    # --------------------------------------------------------

    header(
        "COMPARISON COMPLETE"
    )

    print(
        "Experiments processed:",
        len(
            experiments
        )
    )

    print(
        "Method rows:",
        len(
            rows
        )
    )

    print(
        "\nExcel:"
    )

    print(
        output
    )

    print(
        "\nExcel exists:",
        output.exists()
    )


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":

    main()